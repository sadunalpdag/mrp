import os, time, json, requests, io
from datetime import datetime, timezone

# ========= AYARLAR =========
LIMIT = 300
INTERVALS = ["1h", "4h", "1d"]

# EMA setleri (fast, slow, long)
EMA_SETS = {
    "1h": (7, 25, 99),
    "4h": (9, 26, 200),
    "1d": (20, 50, 200),
}

# ATR & eşikler
ATR_PERIOD = int(os.getenv("ATR_PERIOD", "14"))
ATR_MIN_PCT_DEFAULTS = {"1h": 0.0035, "4h": 0.0025, "1d": 0.0015}
ATR_SLOPE_MULT_DEFAULTS = {"1h": 0.6, "4h": 0.5, "1d": 0.4}

# SL/TP (TP'ler ve raporda 4 ondalık gösterim)
SL_MULT   = float(os.getenv("SL_MULT", "1.5"))
TP1_MULT  = float(os.getenv("TP1_MULT", "1.0"))
TP2_MULT  = float(os.getenv("TP2_MULT", "2.0"))
TP3_MULT  = float(os.getenv("TP3_MULT", "3.0"))

# RSI & Divergence
RSI_PERIOD = int(os.getenv("RSI_PERIOD", "14"))
RSI_SWING_LOOKBACK = int(os.getenv("RSI_SWING_LOOKBACK", "12"))

# Destek/Direnç lookback (bilgi amaçlı)
SR_LOOKBACK = int(os.getenv("SR_LOOKBACK", "100"))

# Erken onay: bar kapanışına ≤30dk kala canlı barda yön korunuyorsa sinyal ver
EARLY_CONFIRM_MS = int(os.getenv("EARLY_CONFIRM_MS", str(30*60*1000)))  # 30dk

# SCALP ayarları (EMA7 slope reversal)
SCALP_TF_CONFIRM = os.getenv("SCALP_TF_CONFIRM", "4h")  # 1h için üst TF onayı
SCALP_TP_MULT = float(os.getenv("SCALP_TP_MULT", "0.5"))
SCALP_SL_MULT = float(os.getenv("SCALP_SL_MULT", "0.25"))
SCALP_MIN_ATR_FACTOR = float(os.getenv("SCALP_MIN_ATR_FACTOR", "1.0"))

# ==== SİMÜLASYON ====
SIM_ENABLE = os.getenv("SIM_ENABLE", "1") == "1"
SIM_MIN_POWER = int(os.getenv("SIM_MIN_POWER", "60"))   # POWER ≥ 60 ise simülasyon al
SIM_TP_PCT = float(os.getenv("SIM_TP_PCT", "0.01"))     # +%1 TP
# SL yüzde: 0.03 / 0.04 / 0.05 (kullanıcı isteğine göre)
SIM_SL_PCT = float(os.getenv("SIM_SL_PCT", "0.1"))
# Rapor aralığı (dakika): varsayılan 60 = saatlik rapor
REPORT_INTERVAL_MIN = int(os.getenv("REPORT_INTERVAL_MIN", "60"))

SLEEP_BETWEEN = 0.2
SCAN_INTERVAL = 300  # 5 dakika

BOT_TOKEN = os.getenv("BOT_TOKEN")
CHAT_ID   = os.getenv("CHAT_ID")

STATE_FILE = os.getenv("STATE_FILE", "alerts.json")
LOG_FILE   = os.getenv("LOG_FILE",   "log.txt")
PREMIUM_LOG_FILE = os.getenv("PREMIUM_LOG_FILE", "premium.log")
# ===========================


def nowiso():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def log(msg):
    print(msg, flush=True)
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"{datetime.now()} - {msg}\n")
    except:
        pass


def log_premium(msg):
    try:
        with open(PREMIUM_LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"{datetime.now()} - {msg}\n")
    except:
        pass


def send_telegram(text):
    if not BOT_TOKEN or not CHAT_ID:
        log("Telegram env eksik (BOT_TOKEN/CHAT_ID). Mesaj gönderilmedi.")
        return
    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        requests.post(url, data={"chat_id": CHAT_ID, "text": text}, timeout=12)
    except Exception as e:
        log(f"Telegram hatası: {e}")


def send_telegram_document(file_bytes: bytes, filename: str, caption: str = ""):
    """Excel/CSV dosyasını Telegram'a yükler."""
    if not BOT_TOKEN or not CHAT_ID:
        log("Telegram env eksik (BOT_TOKEN/CHAT_ID). Belge gönderilmedi.")
        return
    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
        files = {'document': (filename, file_bytes)}
        data = {'chat_id': CHAT_ID, 'caption': caption}
        requests.post(url, files=files, data=data, timeout=20)
    except Exception as e:
        log(f"Telegram sendDocument hatası: {e}")


def safe_load_json(path):
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except:
        pass
    return {}


def safe_save_json(path, data):
    try:
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp, path)
    except:
        pass


# ---------- İNDİKATÖRLER ----------
def ema(values, length):
    ema_vals = [values[0]]
    k = 2 / (length + 1)
    for i in range(1, len(values)):
        ema_vals.append(values[i] * k + ema_vals[-1] * (1 - k))
    return ema_vals

def slope_value(series, lookback=3):
    return series[-1] - series[-lookback] if len(series) > lookback else 0.0

def atr_series(highs, lows, closes, period):
    trs = []
    for i in range(len(highs)):
        if i == 0:
            trs.append(highs[i] - lows[i])
        else:
            pc = closes[i - 1]
            trs.append(max(highs[i] - lows[i], abs(highs[i] - pc), abs(lows[i] - pc)))
    if len(trs) < period:
        return [0.0]*len(trs)
    atr = [sum(trs[:period]) / period]
    for i in range(period, len(trs)):
        atr.append((atr[-1] * (period - 1) + trs[i]) / period)
    return [0.0]*(len(trs)-len(atr)) + atr

def rsi(values, period=14):
    n = len(values)
    if n < period + 1:
        return [None] * n
    deltas = [values[i] - values[i-1] for i in range(1, n)]
    gains  = [max(d, 0.0) for d in deltas]
    losses = [abs(min(d, 0.0)) for d in deltas]
    avg_gain = sum(gains[:period]) / period
    avg_loss = sum(losses[:period]) / period
    rsis = []
    for i in range(period, len(deltas)):
        avg_gain = (avg_gain * (period - 1) + gains[i]) / period
        avg_loss = (avg_loss * (period - 1) + losses[i]) / period
        if avg_loss == 0:
            rsis.append(100.0)
        else:
            rs = avg_gain / avg_loss
            rsis.append(100 - 100 / (1 + rs))
    return [None]*(n-len(rsis)) + rsis

def _local_extrema(series):
    peaks, troughs = [], []
    for i in range(1, len(series)-1):
        if series[i] > series[i-1] and series[i] > series[i+1]:
            peaks.append((i, series[i]))
        if series[i] < series[i-1] and series[i] < series[i+1]:
            troughs.append((i, series[i]))
    return peaks, troughs

def detect_rsi_divergence(closes, rsis, lookback=12):
    if len(closes) < lookback + 3:
        return None, None
    closes = closes[-lookback:]
    rsis   = rsis[-lookback:]
    peaks_p, troughs_p = _local_extrema(closes)
    peaks_r, troughs_r = _local_extrema([r for r in rsis if r is not None])
    if len(peaks_p) >= 2 and len(peaks_r) >= 2:
        if peaks_p[-1][1] > peaks_p[-2][1] and peaks_r[-1][1] < peaks_r[-2][1]:
            return "BEARISH", peaks_p[-1][0]
    if len(troughs_p) >= 2 and len(troughs_r) >= 2:
        if troughs_p[-1][1] < troughs_p[-2][1] and troughs_r[-1][1] > troughs_r[-2][1]:
            return "BULLISH", troughs_p[-1][0]
    return None, None

# --- Stabilizasyon (kapalı bar) + Erken onay (canlı bar) ---
def stabilized_or_early(ema_f_closed, ema_s_closed, ema_f_full, ema_s_full, bar_close_ms, now_ms, early_ms):
    if len(ema_f_closed) < 3:
        return None, None
    prev_diff   = ema_f_closed[-3] - ema_s_closed[-3]
    cross_diff  = ema_f_closed[-2] - ema_s_closed[-2]
    after_diff  = ema_f_closed[-1] - ema_s_closed[-1]
    if prev_diff < 0 and cross_diff > 0 and after_diff > 0:
        return "UP", "CLOSED"
    if prev_diff > 0 and cross_diff < 0 and after_diff < 0:
        return "DOWN", "CLOSED"
    if (bar_close_ms - now_ms) <= early_ms and len(ema_f_full) >= 2 and len(ema_s_full) >= 2:
        curr_live = ema_f_full[-1] - ema_s_full[-1]
        if prev_diff < 0 and cross_diff > 0 and curr_live > 0:
            return "UP", "EARLY"
        if prev_diff > 0 and cross_diff < 0 and curr_live < 0:
            return "DOWN", "EARLY"
    return None, None

def trend_lines_from_extrema(closes, lookback=100):
    clip = closes[-min(lookback, len(closes)):]
    peaks, troughs = _local_extrema(clip)
    if len(peaks) >= 2:
        resistance = (peaks[-1][1] + peaks[-2][1]) / 2.0
    else:
        resistance = max(clip) if clip else None
    if len(troughs) >= 2:
        support = (troughs[-1][1] + troughs[-2][1]) / 2.0
    else:
        support = min(clip) if clip else None
    return support, resistance


# ---------- Binance Kaynak ----------
SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "EMA-ULTRA/2.0", "Accept": "application/json"})

def get_klines(symbol, interval, limit=LIMIT):
    url = "https://fapi.binance.com/fapi/v1/klines"
    params = {"symbol": symbol, "interval": interval, "limit": limit}
    for _ in range(3):
        try:
            r = SESSION.get(url, params=params, timeout=10)
            if r.status_code == 200:
                return r.json()
        except:
            time.sleep(0.4)
    return []

def get_futures_symbols():
    try:
        r = SESSION.get("https://fapi.binance.com/fapi/v1/exchangeInfo", timeout=10)
        data = r.json()
        return [s["symbol"] for s in data["symbols"] if s["quoteAsset"]=="USDT" and s["status"]=="TRADING"]
    except:
        return []


# ---------- ÜST TF YÖN HESABI (1h için 4h & 1d) ----------
def ema_trend_direction_from_closes(closes, fast_len, slow_len):
    ef = ema(closes, fast_len)
    es = ema(closes, slow_len)
    if ef[-1] > es[-1]: return "UP"
    if ef[-1] < es[-1]: return "DOWN"
    return "FLAT"

def get_higher_tf_dirs(sym):
    higher_dirs = {}
    for tf in ["4h", "1d"]:
        kl = get_klines(sym, tf, limit=150)
        if not kl:
            higher_dirs[tf] = "FLAT"; continue
        closes = [float(k[4]) for k in kl]
        ef, es, _ = EMA_SETS[tf]
        higher_dirs[tf] = ema_trend_direction_from_closes(closes, ef, es)
        time.sleep(0.05)
    return higher_dirs

def alignment_score(dirn, higher_dirs):
    score = 0
    matches = sum(1 for d in higher_dirs.values() if d == dirn)
    opposes = sum(1 for d in higher_dirs.values() if (d != dirn and d != "FLAT"))
    if matches == 2: score += 10
    elif matches == 1: score += 5
    if opposes == 2: score -= 10
    return score

def arrow_for(d): return "↑" if d == "UP" else "↓" if d == "DOWN" else "-"


# ---------- EŞİKLER ----------
def thresholds(interval):
    return ATR_MIN_PCT_DEFAULTS.get(interval, 0.003), ATR_SLOPE_MULT_DEFAULTS.get(interval, 0.5)

def sl_tp_from_atr(entry, atr, dirn):
    sl  = entry - SL_MULT * atr if dirn=="UP" else entry + SL_MULT * atr
    tp1 = entry + TP1_MULT * atr if dirn=="UP" else entry - TP1_MULT * atr
    tp2 = entry + TP2_MULT * atr if dirn=="UP" else entry - TP2_MULT * atr
    tp3 = entry + TP3_MULT * atr if dirn=="UP" else entry - TP3_MULT * atr
    risk = abs(entry - sl)
    def rr(tp): return (abs(tp - entry) / risk) if risk > 0 else 0.0
    return sl, tp1, tp2, tp3, rr(tp1), rr(tp2), rr(tp3)


# ---------- Smart Scalp Trigger ----------
def detect_slope_reversal(ema_series):
    if len(ema_series) < 6:
        return None, (0.0, 0.0)
    slope_now  = ema_series[-1] - ema_series[-4]
    slope_prev = ema_series[-2] - ema_series[-5]
    if slope_prev < 0 and slope_now > 0:
        return "UP", (slope_prev, slope_now)
    if slope_prev > 0 and slope_now < 0:
        return "DOWN", (slope_prev, slope_now)
    return None, (slope_prev, slope_now)


# ========== SİMÜLASYON DURUMU ==========
def ensure_sim_state(state):
    if "positions" not in state: state["positions"] = {}   # open positions by symbol
    if "history"   not in state: state["history"]   = []   # closed trades list
    if "last_report_ts" not in state: state["last_report_ts"] = 0
    return state

def open_position(state, symbol, side, price, source, power=None, slope_factor=None, atr_pct=None, sl_pct=SIM_SL_PCT, tp_pct=SIM_TP_PCT):
    if not SIM_ENABLE: return
    pos = state["positions"].get(symbol)
    if pos and pos.get("is_open"):  # aynı sembolde tek pozisyon
        return
    state["positions"][symbol] = {
        "is_open": True,
        "symbol": symbol,
        "side": side,  # LONG/SHORT
        "entry": price,
        "tp_pct": tp_pct,
        "sl_pct": sl_pct,
        "opened_at": nowiso(),
        "opened_ts": int(datetime.now(timezone.utc).timestamp()),
        "bars_held": 0,
        "source": source,            # POWER / SCALP
        "power": power,
        "slope_factor": slope_factor,
        "atr_pct": atr_pct,
    }

def update_positions_and_close_if_hit(state, symbol, last_price):
    """TP/SL kontrolü, kapananları history'ye yazar."""
    pos = state["positions"].get(symbol)
    if not pos or not pos.get("is_open"): return
    side = pos["side"]
    entry = pos["entry"]
    tp_pct = pos["tp_pct"]
    sl_pct = pos["sl_pct"]
    # hedefler
    if side == "LONG":
        tp_price = entry * (1 + tp_pct)
        sl_price = entry * (1 - sl_pct)
        hit_tp = last_price >= tp_price
        hit_sl = last_price <= sl_price
    else:
        tp_price = entry * (1 - tp_pct)
        sl_price = entry * (1 + sl_pct)
        hit_tp = last_price <= tp_price
        hit_sl = last_price >= sl_price

    if hit_tp or hit_sl:
        outcome = "TP" if hit_tp else "SL"
        pnl_pct = tp_pct if hit_tp else -sl_pct
        closed = {
            "symbol": symbol,
            "side": side,
            "entry": round(entry, 6),
            "exit": round(last_price, 6),
            "pnl_pct": round(pnl_pct*100, 2),
            "outcome": outcome,
            "opened_at": pos["opened_at"],
            "closed_at": nowiso(),
            "bars_held": pos.get("bars_held", 0),
            "source": pos.get("source"),
            "power": pos.get("power"),
            "slope_factor": pos.get("slope_factor"),
            "atr_pct": pos.get("atr_pct"),
        }
        state["history"].append(closed)
        # pozisyonu kapat
        state["positions"][symbol] = {"is_open": False}
        # Telegram mini bildirim
        send_telegram(
            f"📘 SIM | {outcome} | {symbol} {side}\n"
            f"Entry: {entry:.6f}  Exit: {last_price:.6f}\n"
            f"PnL: {pnl_pct*100:.2f}%  Bars: {closed['bars_held']}\n"
            f"From: {pos.get('source')} (Power={pos.get('power')})"
        )

def tick_positions_bars_held(state, symbol):
    pos = state["positions"].get(symbol)
    if pos and pos.get("is_open"):
        pos["bars_held"] = pos.get("bars_held", 0) + 1


def make_excel_report_bytes(history, best_candidate):
    """
    .xlsx üretmeye çalışır; olmazsa CSV döner.
    Telegram'a bytes göndeririz.
    """
    # Önce CSV string hazırla (Excel açabilir)
    headers = [
        "symbol","side","entry","exit","pnl_pct","outcome",
        "opened_at","closed_at","bars_held","source","power","slope_factor","atr_pct"
    ]
    import csv
    csv_buf = io.StringIO()
    writer = csv.DictWriter(csv_buf, fieldnames=headers)
    writer.writeheader()
    for row in history:
        writer.writerow(row)
    csv_bytes = csv_buf.getvalue().encode("utf-8")

    # openpyxl varsa xlsx üret
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append(headers)
        for row in history:
            ws.append([row.get(h) for h in headers])
        # ikinci sayfa: Best Candidate
        ws2 = wb.create_sheet("BestCandidate")
        ws2.append(["symbol","interval","slope_factor","atr_pct","note"])
        if best_candidate:
            ws2.append([
                best_candidate.get("symbol"),
                best_candidate.get("interval"),
                round(best_candidate.get("slope_factor",0.0), 3),
                f"{best_candidate.get('atr_pct',0.0)*100:.2f}%",
                "Highest Slope/ATR among valid ATR%"
            ])
        # sütun genişlikleri
        for i in range(1, 14):
            ws.column_dimensions[get_column_letter(i)].width = 14
        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue(), "report.xlsx"
    except Exception as e:
        log(f"openpyxl yok veya xlsx hata: {e}; CSV'ye düşüldü.")
        return csv_bytes, "report.csv"


def maybe_send_periodic_report(state):
    """Belirli aralıklarla kapanan işlemler raporunu gönder."""
    if not state.get("history"):
        return
    now_ts = int(datetime.now(timezone.utc).timestamp())
    last = state.get("last_report_ts", 0)
    if now_ts - last < REPORT_INTERVAL_MIN * 60:
        return

    # Best candidate notunu ekleyelim (son taramadan)
    best = state.get("best_candidate")

    file_bytes, fname = make_excel_report_bytes(state["history"], best)
    caption = f"📊 SIM Raporu | İşlem sayısı: {len(state['history'])}\nTP={int(SIM_TP_PCT*100)}% | SL={int(SIM_SL_PCT*100)}% | POWER≥{SIM_MIN_POWER}"
    send_telegram_document(file_bytes, fname, caption)

    state["last_report_ts"] = now_ts
    # İstersen history'yi sıfırlayabilirsin; şimdilik biriktirelim
    safe_save_json(STATE_FILE, state)


# ---------- ANA İŞ AKIŞI ----------
def process_symbol(sym, state):
    ensure_sim_state(state)

    for interval in INTERVALS:
        kl = get_klines(sym, interval)
        if not kl or len(kl) < 220:
            continue

        closes = [float(k[4]) for k in kl]  # canlı dahil
        highs  = [float(k[2]) for k in kl]
        lows   = [float(k[3]) for k in kl]
        bar_close_ms = int(kl[-1][6])   # canlı barın planlanan kapanışı
        prev_bar_ms  = int(kl[-2][6])   # son kapanmış barın kapanışı
        now_ms = int(datetime.now(timezone.utc).timestamp() * 1000)
        price  = closes[-1]

        ef, es, _ = EMA_SETS[interval]

        # KAPANMIŞ ve TAM seri EMA'ları
        closes_closed = closes[:-1]
        if len(closes_closed) < 3:
            continue
        ema_f_closed = ema(closes_closed, ef)
        ema_s_closed = ema(closes_closed, es)
        ema_f_full   = ema(closes, ef)
        ema_s_full   = ema(closes, es)

        # Stabilizasyon + Erken onay kontrolü (Sadece CROSS için)
        dirn, confirm_mode = stabilized_or_early(
            ema_f_closed, ema_s_closed,
            ema_f_full,   ema_s_full,
            bar_close_ms, now_ms,
            EARLY_CONFIRM_MS
        )

        key  = f"{sym}_{interval}"
        prev = state.get(key, {})

        # ==== 1) ANA CROSS / POWER SİNYALİ ====
        if dirn:
            # Aynı PREV bar için tekrar etme
            if prev.get("last_signal_bar_ms") != prev_bar_ms or prev.get("last_dir") != dirn:
                # ATR / RSI
                atr = atr_series(highs, lows, closes, ATR_PERIOD)
                atr_now = atr[-1]
                atr_pct = (atr_now / price) if price > 0 else 0.0
                slope_now = slope_value(ema_f_closed, 3)
                min_pct, slope_mult = thresholds(interval)
                rsis = rsi(closes, RSI_PERIOD)
                rsi_val = rsis[-1] if rsis[-1] is not None else 50.0
                div_type, _ = detect_rsi_divergence(closes, rsis, RSI_SWING_LOOKBACK)
                rsi_status = f"{div_type} DIVERGENCE" if div_type else "NÖTR"
                support, resistance = trend_lines_from_extrema(closes, lookback=SR_LOOKBACK)

                # ---- Dynamic Momentum Power
                atr_ok = (atr_pct >= min_pct) and (atr_now > 0)
                slope_factor = (abs(slope_now) / (atr_now * slope_mult)) if (atr_now > 0 and slope_mult > 0) else 0.0
                atr_factor = (atr_pct / min_pct) if (min_pct > 0) else 1.0
                momentum_boost = min(25.0, (slope_factor * 10.0) + (atr_factor * 8.0))
                if slope_factor >= 2.0: momentum_tag = "🔥 Aşırı Güçlü"
                elif slope_factor >= 1.0: momentum_tag = "💪 Sağlam"
                elif slope_factor >= 0.5: momentum_tag = "⚠️ Zayıf"
                else: momentum_tag = "🧊 Çok Zayıf"

                power = 40.0 + momentum_boost
                if div_type:
                    power += 10
                    if (dirn == "UP" and div_type == "BULLISH") or (dirn == "DOWN" and div_type == "BEARISH"):
                        power += 15
                if not atr_ok:
                    power -= 3

                trend_line = None
                if interval == "1h":
                    higher_dirs = get_higher_tf_dirs(sym)
                    add = alignment_score(dirn, higher_dirs)
                    power += add
                    indicator = "✅" if add > 0 else "❌" if add < 0 else "➖"
                    trend_line = f"Trend Uyumu: {indicator} 4h{arrow_for(higher_dirs.get('4h','-'))} | 1d{arrow_for(higher_dirs.get('1d','-'))}"

                power = max(0.0, min(power, 100.0))

                # Power etiketi
                if power >= 86:
                    power_tag = "🟦 Ultra Power"; level = "ULTRA"
                elif power >= 70:
                    power_tag = "🟩 Strong";      level = "PREMIUM"
                elif power >= 50:
                    power_tag = "🟨 Moderate";    level = "CROSS"
                else:
                    power_tag = "🟥 Weak";        level = "CROSS"

                # SL/TP (bilgi)
                sl, tp1, tp2, tp3, rr1, rr2, rr3 = sl_tp_from_atr(price, atr_now, dirn)
                rr = lambda x: f"{x:.2f}"
                atr_tag = "[ATR OK]" if atr_ok else "[ATR LOW]"
                early_tag = " ⏳ Early Confirm (30dk)" if confirm_mode == "EARLY" else ""
                tag = "⚡ CROSS"
                if level == "PREMIUM": tag = "⚡🔥 PREMIUM SİNYAL"
                elif level == "ULTRA": tag = "⚡🚀 ULTRA PREMIUM SİNYAL"

                # ---- Telegram
                lines = [
                    f"{tag}: {sym} ({interval}) {atr_tag}{early_tag}",
                    f"Power: {power_tag} ({power:.0f}/100)",
                    f"Momentum: {momentum_tag} (Slope/ATR={slope_factor:.2f}, ATR%={atr_pct*100:.2f}%)",
                    f"Direction: {dirn} ({'LONG' if dirn=='UP' else 'SHORT'})",
                    f"Kesişim: EMA{ef} {'↗' if dirn=='UP' else '↘'} EMA{es}",
                    trend_line,
                    f"RSI: {rsi_val:.2f} → {rsi_status}",
                    f"ATR({ATR_PERIOD}): {atr_now:.6f} ({atr_pct*100:.2f}%)",
                    f"Slope(fast,3): {slope_now:.6f}",
                    f"Support≈ {support:.4f} | Resistance≈ {resistance:.4f}" if (support is not None and resistance is not None) else None,
                    f"Eşik[{interval}]: ATR%≥{min_pct*100:.2f} | slope_mult={slope_mult:.2f}",
                    f"Entry≈ {price:.6f}",
                    f"SL≈ {sl:.6f} | TP1≈ {tp1:.4f} (R:R {rr(rr1)})  TP2≈ {tp2:.4f} (R:R {rr(rr2)})  TP3≈ {tp3:.4f} (R:R {rr(rr3)})",
                    f"Time: {nowiso()}",
                ]
                msg = "\n".join([l for l in lines if l])
                send_telegram(msg)
                if level in ("PREMIUM", "ULTRA"):
                    log_premium(msg)

                # ==== SIMÜLASYON GİRİŞ (POWER ≥ SIM_MIN_POWER) ====
                if SIM_ENABLE and power >= SIM_MIN_POWER and interval == "1h":
                    side = "LONG" if dirn == "UP" else "SHORT"
                    open_position(
                        state, sym, side, price, source="POWER",
                        power=round(power, 0),
                        slope_factor=round(slope_factor, 3),
                        atr_pct=atr_pct,
                        sl_pct=SIM_SL_PCT, tp_pct=SIM_TP_PCT
                    )

                # Sinyali stabilizasyon barına pin'le → kapanışta tekrar etmez
                state[key] = {**prev, "last_signal_bar_ms": prev_bar_ms, "last_dir": dirn}
                safe_save_json(STATE_FILE, state)
                time.sleep(SLEEP_BETWEEN)

        # ==== 2) SMART SCALP TRIGGER (bağımsız ve her zaman SIM alır) ====
        if interval == "1h":
            slope_flip, (s_prev, s_now) = detect_slope_reversal(ema_f_closed)
            if slope_flip:
                if prev.get("last_scalp_bar_ms") != prev_bar_ms or prev.get("last_scalp_dir") != slope_flip:
                    higher_dirs = get_higher_tf_dirs(sym)
                    tf_dir = higher_dirs.get(SCALP_TF_CONFIRM, "FLAT")

                    atr = atr_series(highs, lows, closes, ATR_PERIOD)
                    atr_now = atr[-1]
                    atr_pct = (atr_now / price) if price > 0 else 0.0
                    min_pct_1h = ATR_MIN_PCT_DEFAULTS["1h"]
                    atr_ok_for_scalp = atr_pct >= (min_pct_1h * SCALP_MIN_ATR_FACTOR)

                    if slope_flip == tf_dir and atr_ok_for_scalp:
                        tp = price + (SCALP_TP_MULT * atr_now if slope_flip == "UP" else -SCALP_TP_MULT * atr_now)
                        sl = price - (SCALP_SL_MULT * atr_now if slope_flip == "UP" else -SCALP_SL_MULT * atr_now)

                        denom = (atr_now * ATR_SLOPE_MULT_DEFAULTS["1h"]) if atr_now > 0 else 1.0
                        scalp_power = max(0.0, min(100.0, 60.0 + (abs(s_now - s_prev) / denom) * 20.0))

                        scalp_text = (
                            f"💥 SCALP {('LONG' if slope_flip=='UP' else 'SHORT')} TRIGGER: {sym} (1h)\n"
                            f"{SCALP_TF_CONFIRM.upper()} Trend Onayı: {tf_dir} ✅\n"
                            f"Slope Change: {s_prev:+.6f} → {s_now:+.6f}\n"
                            f"ATR({ATR_PERIOD}): {atr_now:.6f} ({atr_pct*100:.2f}%)\n"
                            f"TP≈ {tp:.4f} | SL≈ {sl:.4f}  (TP {SCALP_TP_MULT}×ATR, SL {SCALP_SL_MULT}×ATR)\n"
                            f"Power: ⚡ Momentum {scalp_power:.0f}\n"
                            f"Time: {nowiso()}"
                        )
                        send_telegram(scalp_text)

                        # ==== SIMÜLASYON GİRİŞ (SCALP) ====
                        side = "LONG" if slope_flip == "UP" else "SHORT"
                        if SIM_ENABLE:
                            open_position(
                                state, sym, side, price, source="SCALP",
                                power=round(scalp_power, 0),
                                slope_factor=None, atr_pct=atr_pct,
                                sl_pct=SIM_SL_PCT, tp_pct=SIM_TP_PCT
                            )

                        state[key] = {**state.get(key, {}), "last_scalp_bar_ms": prev_bar_ms, "last_scalp_dir": slope_flip}
                        safe_save_json(STATE_FILE, state)
                        time.sleep(SLEEP_BETWEEN)

        # ==== 3) SİMÜLASYON POZİSYON GÜNCELLEME ====
        # her sembol için bar kapanışında bar sayacı 1 artar ve TP/SL kontrol edilir
        tick_positions_bars_held(state, sym)
        update_positions_and_close_if_hit(state, sym, price)

        # ==== 4) “En Uygun Slope/ATR Dönüş” Adayı (bilgi & rapor için state'e yaz) ====
        if interval == "1h":
            # anlık momentum adayı: ATR yeterli ve slope_factor en yüksek
            atr = atr_series(highs, lows, closes, ATR_PERIOD)
            atr_now = atr[-1]
            atr_pct = (atr_now / price) if price > 0 else 0.0
            min_pct, slope_mult = thresholds("1h")
            if atr_now > 0 and atr_pct >= min_pct:
                ef1, es1, _ = EMA_SETS["1h"]
                emaf = ema(closes, ef1)[:-1]  # kapalı barda momentum
                slope_now = slope_value(emaf, 3)
                slope_factor = abs(slope_now) / (atr_now * slope_mult) if slope_mult > 0 else 0.0
                cur_best = state.get("best_candidate")
                if (not cur_best) or slope_factor > cur_best.get("slope_factor", 0.0):
                    state["best_candidate"] = {
                        "symbol": sym,
                        "interval": "1h",
                        "slope_factor": slope_factor,
                        "atr_pct": atr_pct,
                        "price": price,
                        "time": nowiso(),
                    }
                    safe_save_json(STATE_FILE, state)


def main():
    log("🚀 v9 | Binance only | EMA/ATR/RSI — Power + Smart Scalp + SIM(+1% TP, −3/4/5% SL) + Excel Rapor")
    state = safe_load_json(STATE_FILE)
    ensure_sim_state(state)

    binance_symbols = get_futures_symbols()
    if not binance_symbols:
        log("❌ Binance sembol listesi boş.")
        return

    while True:
        for sym in binance_symbols:
            process_symbol(sym, state)
        # periyodik rapor
        maybe_send_periodic_report(state)

        log("⏳ 5 dk bekleniyor...\n")
        time.sleep(SCAN_INTERVAL)


if __name__ == "__main__":
    main()